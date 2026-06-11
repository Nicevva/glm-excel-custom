"""Create 4 demo Excel files for GIF recording."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, cols, min_width=12):
    for c in range(1, cols + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 4)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len, 28)


# ──────────────────────────────────────────────
# 1. Formula Generation Demo
# ──────────────────────────────────────────────
def create_formula_demo():
    wb = openpyxl.Workbook()

    # Sheet 1: Sales data
    ws = wb.active
    ws.title = "Sales Data"
    headers = ["Order ID", "Date", "Region", "Product", "Category", "Qty", "Unit Price", "Discount", "Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    sales = [
        ["ORD-001", "2026-01-05", "East",  "Laptop Pro 15",    "Electronics", 3,  6999, 0.05, None],
        ["ORD-002", "2026-01-08", "South", "Wireless Mouse",   "Accessories", 25, 89,   0,    None],
        ["ORD-003", "2026-01-12", "East",  "Mechanical KB",    "Accessories", 10, 359,  0.1,  None],
        ["ORD-004", "2026-01-15", "North", "Monitor 27in",     "Electronics", 5,  2499, 0.08, None],
        ["ORD-005", "2026-01-20", "West",  "USB-C Hub",        "Accessories", 50, 129,  0.15, None],
        ["ORD-006", "2026-02-03", "East",  "Laptop Pro 15",    "Electronics", 2,  6999, 0.05, None],
        ["ORD-007", "2026-02-10", "South", "Webcam HD",        "Accessories", 15, 299,  0,    None],
        ["ORD-008", "2026-02-14", "North", "Laptop Air 13",    "Electronics", 8,  5299, 0.1,  None],
        ["ORD-009", "2026-02-28", "West",  "Wireless Mouse",   "Accessories", 30, 89,   0.05, None],
        ["ORD-010", "2026-03-05", "East",  "Monitor 27in",     "Electronics", 4,  2499, 0,    None],
        ["ORD-011", "2026-03-11", "South", "Mechanical KB",    "Accessories", 20, 359,  0.12, None],
        ["ORD-012", "2026-03-18", "North", "Laptop Pro 15",    "Electronics", 6,  6999, 0.08, None],
        ["ORD-013", "2026-03-25", "West",  "USB-C Hub",        "Accessories", 40, 129,  0.1,  None],
        ["ORD-014", "2026-04-02", "East",  "Webcam HD",        "Accessories", 12, 299,  0.05, None],
        ["ORD-015", "2026-04-10", "South", "Laptop Air 13",    "Electronics", 3,  5299, 0.15, None],
    ]
    for r, row_data in enumerate(sales, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data(ws, 2, len(sales) + 1, len(headers))
    auto_width(ws, len(headers))

    # Sheet 2: Tasks for user
    ws2 = wb.create_sheet("AI Formula Tasks")
    task_headers = ["Task #", "Description (ask AI)", "Target Cell", "Expected Result"]
    for i, h in enumerate(task_headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(task_headers))

    tasks = [
        [1, "Calculate Total = Qty * Unit Price * (1 - Discount)", "I2:I16", "Auto-fill formula"],
        [2, "Total sales for East region", "K2", "=SUMPRODUCT(...)"],
        [3, "Avg discount by category", "K5", "=AVERAGEIFS(...)"],
        [4, "Find the most expensive product", "K8", "=INDEX/MATCH"],
        [5, "Count orders over 10,000 in total", "K11", "=COUNTIF(...)"],
    ]
    for r, row_data in enumerate(tasks, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
    style_data(ws2, 2, len(tasks) + 1, len(task_headers))
    auto_width(ws2, len(task_headers))

    wb.save(os.path.join(OUTPUT_DIR, "demo1_formula_generation.xlsx"))
    print("Created: demo1_formula_generation.xlsx")


# ──────────────────────────────────────────────
# 2. Data Cleaning Demo
# ──────────────────────────────────────────────
def create_cleaning_demo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dirty Data"

    headers = ["Name", "Phone", "Email", "City", "Join Date", "Department", "Salary"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    dirty_data = [
        ["ZHANG SAN",       "13812345678",    "zhangsan@gmail.com",     "Beijing",    "2024-01-15",   "Sales",        15000],
        ["zhang san",       "138-1234-5678",  "zhangsan@gmail.com",     "beijing",    "2024/01/15",   "sales",        15000],
        ["Li Si",           "139 8765 4321",  "lisi@OUTLOOK.COM",       "ShangHai",   "23-Mar-2024",  "Engineering",  22000],
        ["li si ",          "13987654321",    "LISI@outlook.com",       "shanghai",   "2024.3.23",    "engineering",  22000],
        ["Wang Wu",         "15012349876",    "wangwu@company.cn",      "Guang Zhou", "2024-05-01",   "Marketing",    18000],
        ["WANG WU",         "150-1234-9876",  "wangwu@company.cn",      "guangzhou",  "May 1, 2024",  "marketing ",   18000],
        ["Zhao Liu",        "18676543210",    "zhaoliu@qq.com",         "Shenzhen",   "2024-07-20",   "Sales",        16500],
        ["Chen Qi",         "177 0098 1234",  "chenqi@163.com",         "HangZhou",   "2024/09/10",   "Engineering",  25000],
        ["  Chen Qi",       "17700981234",    "CHENQI@163.COM",         "hang zhou",  "2024.09.10",   " Engineering", 25000],
        ["Sun Ba",          "13344556677",    "sunba@@company.cn",      "Chengdu",    "2024-11-01",   "HR",           14000],
        ["ZHOU JIU",        "155-6677-8899",  "zhoujiu@gmail.com",      "Wuhan",      "12/15/2024",   "Finance",      20000],
        ["zhou jiu",        "15566778899",    "zhoujiu@gmail.com ",     "wuhan",      "2024-12-15",   "finance",      20000],
        ["Wu Shi",          "N/A",            "wushi@company",          "Nanjing",    "2025-01-08",   "Sales",        17000],
        ["  Zheng shiyi  ", "18811223344",    "zheng11@outlook.com",    "Xi'an",      "Jan 20 2025",  "Marketing",    19000],
        ["Zheng Shiyi",     "188-1122-3344",  "zheng11@outlook.com",    "xian",       "2025/1/20",    "Marketing",    19000],
    ]

    for r, row_data in enumerate(dirty_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    style_data(ws, 2, len(dirty_data) + 1, len(headers))
    auto_width(ws, len(headers))

    # Highlight issues
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    issue_cells = [
        (3, 1), (3, 2), (3, 4), (3, 5), (3, 6),  # zhang san row
        (4, 1), (4, 2), (4, 4), (4, 5), (4, 6),   # li si row
        (5, 1), (5, 2), (5, 4), (5, 5), (5, 6),   # Li Si dup
        (6, 4),                                     # Guang Zhou
        (7, 1), (7, 2), (7, 4), (7, 5), (7, 6),   # WANG WU row
        (10, 1), (10, 2), (10, 4), (10, 6),        # Chen Qi dup
        (11, 3),                                    # sunba@@ double @
        (13, 1), (13, 2), (13, 4), (13, 5), (13, 6),
        (14, 2),                                    # N/A phone
        (14, 3),                                    # missing domain
    ]
    for r, c in issue_cells:
        ws.cell(row=r, column=c).fill = yellow

    # Sheet 2: Cleaning tasks
    ws2 = wb.create_sheet("Cleaning Tasks")
    task_headers = ["Task #", "Issue", "Ask AI"]
    for i, h in enumerate(task_headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(task_headers))

    tasks = [
        [1, "Name casing inconsistent", "Standardize all names to Title Case"],
        [2, "Phone format varies", "Unify to 13812345678 (no separators)"],
        [3, "Duplicate rows", "Find and highlight duplicate entries"],
        [4, "Date format chaos", "Convert all dates to YYYY-MM-DD"],
        [5, "City name inconsistent", "Standardize city names (e.g. Guang Zhou -> Guangzhou)"],
        [6, "Invalid email", "Flag emails with format errors"],
    ]
    for r, row_data in enumerate(tasks, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
    style_data(ws2, 2, len(tasks) + 1, len(task_headers))
    auto_width(ws2, len(task_headers))

    wb.save(os.path.join(OUTPUT_DIR, "demo2_data_cleaning.xlsx"))
    print("Created: demo2_data_cleaning.xlsx")


# ──────────────────────────────────────────────
# 3. Multi-Model Switching Demo
# ──────────────────────────────────────────────
def create_multi_model_demo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model Comparison"

    headers = ["Task", "Input Data", "GPT-4o Result", "Claude Result", "GLM-4 Result", "Best"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    data = [
        [
            "Sentiment Analysis",
            "The product quality is decent but delivery was terrible, took 2 weeks.",
            "", "", "", ""
        ],
        [
            "Text Summarization",
            "Q3 revenue grew 15% YoY to $2.3B, driven by cloud services (+32%) offsetting hardware decline (-8%). Operating margin expanded 200bps to 18.5%.",
            "", "", "", ""
        ],
        [
            "Formula Generation",
            "Calculate year-over-year growth rate for each quarter",
            "", "", "", ""
        ],
        [
            "Data Classification",
            "Classify these expenses: Coffee $5, AWS $2000, Taxi $30, Slack $12, Lunch $15",
            "", "", "", ""
        ],
        [
            "Translation",
            "The quarterly earnings call highlighted strong momentum in AI-driven revenue streams.",
            "", "", "", ""
        ],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    style_data(ws, 2, len(data) + 1, len(headers))
    auto_width(ws, len(headers))

    # Sheet 2: Quarterly data for formula task
    ws2 = wb.create_sheet("Revenue Data")
    rev_headers = ["Quarter", "2024 Revenue", "2025 Revenue", "YoY Growth"]
    for i, h in enumerate(rev_headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(rev_headers))

    rev_data = [
        ["Q1", 1850000, 2130000, None],
        ["Q2", 2010000, 2350000, None],
        ["Q3", 1920000, 2280000, None],
        ["Q4", 2200000, 2580000, None],
    ]
    for r, row_data in enumerate(rev_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c in (2, 3) and val:
                cell.number_format = '#,##0'
    style_data(ws2, 2, len(rev_data) + 1, len(rev_headers))
    auto_width(ws2, len(rev_headers))

    wb.save(os.path.join(OUTPUT_DIR, "demo3_multi_model.xlsx"))
    print("Created: demo3_multi_model.xlsx")


# ──────────────────────────────────────────────
# 4. Local Privacy Demo
# ──────────────────────────────────────────────
def create_privacy_demo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Data"

    headers = ["Emp ID", "Name", "ID Number", "Phone", "Department", "Position", "Base Salary", "Bonus", "Total Comp"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    # Fake but realistic-looking sensitive data
    emp_data = [
        ["E001", "Zhang Wei",    "110101199001011234", "13812345678", "Engineering", "Senior Dev",    28000, 8400,  None],
        ["E002", "Li Na",        "310115198805202345", "13998765432", "Engineering", "Tech Lead",     35000, 14000, None],
        ["E003", "Wang Fang",    "440305199203153456", "15012341234", "Product",     "Product Mgr",   30000, 9000,  None],
        ["E004", "Liu Yang",     "510107199106084567", "18676541234", "Sales",       "Sales Dir",     32000, 19200, None],
        ["E005", "Chen Jie",     "330106198712125678", "17712345678", "Finance",     "CFO",           45000, 22500, None],
        ["E006", "Zhao Min",     "320102199404016789", "15598765432", "HR",          "HR Manager",    25000, 7500,  None],
        ["E007", "Huang Lei",    "420111199508237890", "13344556677", "Engineering", "Junior Dev",    18000, 3600,  None],
        ["E008", "Zhou Ting",    "500103199307148901", "18811223344", "Marketing",   "Marketing Mgr", 27000, 10800, None],
        ["E009", "Wu Gang",      "610102199009029012", "13655667788", "Engineering", "Senior Dev",    29000, 8700,  None],
        ["E010", "Sun Li",       "210102199211180123", "15766778899", "Sales",       "Sales Rep",     15000, 6000,  None],
    ]

    for r, row_data in enumerate(emp_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (7, 8) and val:
                cell.number_format = '#,##0'
    style_data(ws, 2, len(emp_data) + 1, len(headers))
    auto_width(ws, len(headers))

    # Sensitive highlight
    red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    for r in range(2, len(emp_data) + 2):
        for c in [3, 4, 7, 8]:  # ID, Phone, Salary, Bonus
            ws.cell(row=r, column=c).fill = red_fill

    # Sheet 2: AI analysis tasks
    ws2 = wb.create_sheet("AI Tasks (Local Only)")
    task_headers = ["Task #", "Description", "Why Local Matters"]
    for i, h in enumerate(task_headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(task_headers))

    tasks = [
        [1, "Calculate Total Compensation (Salary + Bonus)", "Salary data must not leave local network"],
        [2, "Analyze salary distribution by department", "Prevents pay data leaks"],
        [3, "Flag employees with bonus > 30% of salary", "Internal compensation policy is confidential"],
        [4, "Generate department headcount & cost summary", "Org structure is sensitive business intel"],
    ]
    for r, row_data in enumerate(tasks, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
    style_data(ws2, 2, len(tasks) + 1, len(task_headers))
    auto_width(ws2, len(task_headers))

    wb.save(os.path.join(OUTPUT_DIR, "demo4_local_privacy.xlsx"))
    print("Created: demo4_local_privacy.xlsx")


if __name__ == "__main__":
    create_formula_demo()
    create_cleaning_demo()
    create_multi_model_demo()
    create_privacy_demo()
    print("\nAll 4 demo Excel files created!")
